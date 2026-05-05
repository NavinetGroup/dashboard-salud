# -*- coding: utf-8 -*-
"""
dift18_scraper.py — Extract DIFT18 Desempeño ET 2024 from Excel.

Source:  data/raw/DIFT18 Tablero de control Desempeño ET 2024 Consolidado *.xlsx
Output:  data/parquet/dift18_departamento.parquet  (32 depts + 6 distritos)
         data/parquet/dift18_municipio.parquet

Sheet layout (confirmed by inspection):
  5_Result_Depart  — row 9+, col 1=nro, col 2=entity name, cols 3-78=38 pairs, col 79=extra_res
  6_Result_Dist    — row 9+, same layout as 5_Result_Depart but 78 cols (no extra_res)
                     6 special districts: Barranquilla, Bogotá D.C., Buenaventura,
                     Cali, Cartagena, Santa Marta  (same 38-indicator structure)
  7_Result_Munic   — row 10+, cols 1-6=geo, cols 7-42=18 pairs, col 43=trailing None
"""

from pathlib import Path

import openpyxl
import polars as pl

BASE_DIR = Path(__file__).resolve().parent.parent
RAW_DIR  = BASE_DIR / 'data' / 'raw'
OUT_DIR  = BASE_DIR / 'data' / 'parquet'

EXCEL_GLOB = 'DIFT18 Tablero de control Desempeño ET 2024 Consolidado*.xlsx'

# ---------------------------------------------------------------------------
# Column schemas
# ---------------------------------------------------------------------------

_DEPT_INDICATORS = [
    'inspva', 'calidad_gaudi', 'cobertura', 'habilitacion',
    'visitas_ips', 'visitas_previas', 'pamec', 'pamec_ips',
    'reporte_sic', 'reporte_st004', 'ips_sic',
    'reporte_525', 'reporte_526', 'lvr_crue',
    'letalidad_dengue', 'muerte_fa', 'mort_menores5', 'mort_materna',
    'tasa_suicidio',
    'reporte_info', 'cuipo', 'equilibrio_fls',
    'ejec_ingresos', 'ejec_compromisos', 'ejec_obligaciones', 'ejec_pagos',
    'sgp_oferta_comp', 'sgp_oferta_oblig', 'sgp_sp_comp', 'sgp_sp_oblig',
    'pagos_noupcc', 'facturacion_pna', 'pagos_pna',
    'facturacion_migrantes', 'pagos_migrantes', 'deuda_esfuerzo',
    'reporte_sns_juegos', 'inoportunidad_rentas',
]  # 38 indicators -> 76 data cols (3-78) + col 79 = extra_res

_MUN_INDICATORS = [
    'cobertura', 'listado_censal', 'sisben',
    'mort_materna', 'mort_menores5', 'desnutricion',
    'letalidad_dengue',
    'reporte_info', 'cuipo', 'equilibrio_fls',
    'ejec_ingresos', 'ejec_compromisos', 'ejec_obligaciones', 'ejec_pagos',
    'sgp_subsidiado', 'sgp_sp_comp', 'sgp_sp_oblig',
    'reporte_sns_juegos',
]  # 18 indicators -> 36 data cols (7-42); col 43 trailing, ignored

_DEPT_COLS = (
    ['nro', 'entidad_territorial']
    + [f'{ind}_{sfx}' for ind in _DEPT_INDICATORS for sfx in ('res', 'pts')]
    + ['extra_res']
)  # 2 + 76 + 1 = 79

# District (6_Result_Dist) has 78 cols — same as dept but no extra_res at the end.
# nro order in that sheet maps to DANE municipality codes for the primary city.
_DIST_MUN_BY_NRO: dict[int, int] = {
    1: 8001,   # BARRANQUILLA
    2: 11001,  # BOGOTÁ D.C.
    3: 76109,  # BUENAVENTURA
    4: 76001,  # CALI
    5: 13001,  # CARTAGENA
    6: 47001,  # SANTA MARTA
}

_MUN_COLS = (
    ['nro', 'direccion_regional', 'codigo_dep', 'departamento', 'codigo_mun', 'municipio']
    + [f'{ind}_{sfx}' for ind in _MUN_INDICATORS for sfx in ('res', 'pts')]
)  # 6 + 36 = 42  (col 43 dropped)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _to_float(v) -> float | None:
    """Convert Excel cell value (may be string '4' or numeric) to float."""
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _puntaje_cols(indicators: list[str]) -> list[str]:
    return [f'{ind}_pts' for ind in indicators]


def _read_dept_sheet(ws) -> pl.DataFrame:
    rows = []
    for raw in ws.iter_rows(min_row=9, values_only=True):
        if raw[1] is None:
            continue
        # Pad or truncate to exactly 79 values
        row = list(raw) + [None] * max(0, 79 - len(raw))
        rows.append(row[:79])

    # Build as list-of-lists per column — force all to Utf8 then cast
    data: dict[str, list] = {col: [] for col in _DEPT_COLS}
    for row in rows:
        for i, col in enumerate(_DEPT_COLS):
            v = row[i]
            data[col].append(None if v is None else str(v))

    schema = {col: pl.Utf8 for col in _DEPT_COLS}
    df = pl.DataFrame(data, schema=schema)

    # Cast all numeric cols (res and pts) to Float64
    num_cols = (
        [f'{ind}_{sfx}' for ind in _DEPT_INDICATORS for sfx in ('res', 'pts')]
        + ['extra_res']
    )
    for col in num_cols:
        if col in df.columns:
            df = df.with_columns(
                pl.col(col).cast(pl.Float64, strict=False).alias(col)
            )

    # Compute average puntaje (1–4 scale)
    pts_cols = _puntaje_cols(_DEPT_INDICATORS)
    pts_cols = [c for c in pts_cols if c in df.columns]
    df = df.with_columns(
        pl.concat_list([pl.col(c) for c in pts_cols])
          .list.drop_nulls()
          .list.mean()
          .alias('puntaje_promedio')
    )

    return df


def _read_mun_sheet(ws) -> pl.DataFrame:
    rows = []
    for raw in ws.iter_rows(min_row=10, values_only=True):
        if raw[1] is None:
            continue
        row = list(raw) + [None] * max(0, 43 - len(raw))
        rows.append(row[:42])  # drop trailing col 43

    data: dict[str, list] = {col: [] for col in _MUN_COLS}
    for row in rows:
        for i, col in enumerate(_MUN_COLS):
            v = row[i]
            data[col].append(None if v is None else str(v))

    schema = {col: pl.Utf8 for col in _MUN_COLS}
    df = pl.DataFrame(data, schema=schema)

    # Cast codes and numeric cols
    df = df.with_columns(
        pl.col('codigo_mun').str.strip_chars().cast(pl.Int32, strict=False),
        pl.col('codigo_dep').str.strip_chars().cast(pl.Int32, strict=False),
    )
    num_cols = [f'{ind}_{sfx}' for ind in _MUN_INDICATORS for sfx in ('res', 'pts')]
    for col in num_cols:
        if col in df.columns:
            df = df.with_columns(
                pl.col(col).cast(pl.Float64, strict=False).alias(col)
            )

    pts_cols = _puntaje_cols(_MUN_INDICATORS)
    pts_cols = [c for c in pts_cols if c in df.columns]
    df = df.with_columns(
        pl.concat_list([pl.col(c) for c in pts_cols])
          .list.drop_nulls()
          .list.mean()
          .alias('puntaje_promedio')
    )

    return df


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def run() -> None:
    files = sorted(RAW_DIR.glob(EXCEL_GLOB))
    if not files:
        print(f'dift18: no Excel file found matching "{EXCEL_GLOB}" in {RAW_DIR}')
        return

    excel_path = files[-1]  # latest if multiple
    print(f'dift18: leyendo {excel_path.name}')

    wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    # Department sheet (32 departments)
    df_dept = _read_dept_sheet(wb['5_Result_Depart'])
    df_dept = df_dept.with_columns(
        pl.lit('departamento').alias('tipo'),
        pl.lit(None).cast(pl.Int32).alias('mun_codigo_ref'),
    )

    # District sheet (6 special cities — same 38-indicator structure)
    df_dist = _read_dept_sheet(wb['6_Result_Dist'])
    df_dist = df_dist.with_columns(
        pl.lit('distrito').alias('tipo'),
        pl.col('nro').cast(pl.Int32, strict=False)
          .map_elements(lambda n: _DIST_MUN_BY_NRO.get(n), return_dtype=pl.Int32)
          .alias('mun_codigo_ref'),
    )

    df_dept_all = pl.concat([df_dept, df_dist], how='diagonal_relaxed')
    out_dept = OUT_DIR / 'dift18_departamento.parquet'
    df_dept_all.write_parquet(out_dept, compression='zstd')
    print(f'  dift18_departamento: {len(df_dept)} depts + {len(df_dist)} distritos = {len(df_dept_all)} filas -> {out_dept.name}')

    # Municipality sheet
    df_mun = _read_mun_sheet(wb['7_Result_Munic'])
    out_mun = OUT_DIR / 'dift18_municipio.parquet'
    df_mun.write_parquet(out_mun, compression='zstd')
    print(f'  dift18_municipio: {len(df_mun)} filas -> {out_mun.name}')

    wb.close()
    print('dift18: extracción completa.')


if __name__ == '__main__':
    run()
