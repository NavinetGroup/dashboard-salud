# -*- coding: utf-8 -*-
"""
Scraper afiliación en salud — MinSalud.

Descarga el ZIP mensual publicado en MinSalud, extrae el Excel interior,
lee la hoja CoberturaMunicipio y escribe directamente a Parquet + DuckDB.
No produce archivos CSV intermedios.
"""

import io
import logging
import re
import time
import zipfile
from datetime import datetime
from io import BytesIO
from pathlib import Path

import duckdb
import openpyxl
import polars as pl
import requests

log = logging.getLogger(__name__)

BASE_URL = (
    'https://www.minsalud.gov.co/sites/rid/Lists/BibliotecaDigital/RIDE/VP/RBC/'
    'cifras-afiliacion-salud-{mes}-{year}.zip'
)

# Known historical URLs that don't match the standard pattern
HISTORICAL_URLS: dict[tuple[int, int], list[str]] = {
    (2025, 7):  ['https://www.minsalud.gov.co/sites/rid/Lists/BibliotecaDigital/RIDE/VP/DOA/OAS/cifras-afiliacion-salud-julio-2025.zip'],
    (2025, 8):  ['https://www.minsalud.gov.co/sites/rid/Lists/BibliotecaDigital/RIDE/VP/DOA/OAS/cifras-afiliacion-salud-agosto-2025.zip'],
    (2025, 9):  ['https://www.minsalud.gov.co/sites/rid/Lists/BibliotecaDigital/RIDE/VP/DOA/OAS/cifras-afiliacion-septiembre-2025.zip'],
    (2025, 10): ['https://www.minsalud.gov.co/sites/rid/Lists/BibliotecaDigital/RIDE/VP/DOA/OAS/cifras-afiliacion-salud-oct2025.zip'],
    (2025, 11): ['https://www.minsalud.gov.co/sites/rid/Lists/BibliotecaDigital/RIDE/VP/DOA/cifras-afiliacion-salud-nov-2025.zip'],
    (2025, 12): ['https://www.minsalud.gov.co/sites/rid/Lists/BibliotecaDigital/RIDE/VP/DOA/cifras-afiliacion-salud-dic-2025.zip'],
    (2026, 1):  ['https://www.minsalud.gov.co/sites/rid/Lists/BibliotecaDigital/RIDE/VP/RBC/cifra-afiliacion-salud-enero-2026.zip'],
    (2026, 2):  ['https://www.minsalud.gov.co/sites/rid/Lists/BibliotecaDigital/RIDE/VP/RBC/cifras-afiliacion-salud-feb-2026.zip'],
    (2026, 3):  ['https://www.minsalud.gov.co/sites/rid/Lists/BibliotecaDigital/RIDE/VP/RBC/cifras-afiliacion-salud-marzo-2026.zip'],
}

MESES_ES = {
    1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril',
    5: 'mayo', 6: 'junio', 7: 'julio', 8: 'agosto',
    9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre',
}

RENAME_MAP = {
    'IdDepartamento': 'codigo_dep',
    'Departamento': 'departamento',
    'IdMunicipio': 'codigo_mun',
    'Municipio': 'municipio',
    'Contributivo': 'afiliados_contributivo',
    'Subsidiado': 'afiliados_subsidiado',
    'Total': 'afiliados_total',
    'DEPARTAMENTO': 'departamento',
    'MUNICIPIO': 'municipio',
    'REGIMEN': 'regimen',
    'Régimen': 'regimen',
    'ENTIDAD': 'entidad',
    'Entidad': 'entidad',
    'AFILIADOS': 'afiliados_total',
    'Afiliados': 'afiliados_total',
    'TOTAL AFILIADOS': 'afiliados_total',
    'Total Afiliados': 'afiliados_total',
    'CODIGO DEPARTAMENTO': 'codigo_dep',
    'CODIGO MUNICIPIO': 'codigo_mun',
    'Código Departamento': 'codigo_dep',
    'Código Municipio': 'codigo_mun',
}


def _build_url(year: int, month: int) -> str:
    historical = HISTORICAL_URLS.get((year, month))
    if historical:
        return historical[0]
    return BASE_URL.format(mes=MESES_ES[month], year=year)


def _download_zip(url: str, retries: int = 3) -> bytes | None:
    headers = {'User-Agent': 'Mozilla/5.0'}
    for attempt in range(retries):
        try:
            r = requests.get(url, headers=headers, timeout=(10, 300), stream=True)
            if r.status_code != 200:
                log.warning(f'afiliacion: HTTP {r.status_code} para {url}')
                return None
            chunks = [c for c in r.iter_content(1024 * 1024) if c]
            return b''.join(chunks)
        except Exception as e:
            if attempt < retries - 1:
                wait = 20 * (attempt + 1)
                log.warning(f'afiliacion: intento {attempt + 1} fallido, reintentando en {wait}s — {e}')
                time.sleep(wait)
            else:
                log.warning(f'afiliacion: {retries} intentos fallidos para {url} — {e}')
    return None


def _read_sheet_openpyxl(xlsx_bytes: bytes, sheet_name: str) -> list | None:
    """Read one sheet of an Excel file as a list of row tuples using openpyxl."""
    try:
        wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return None
        ws = wb[sheet_name]
        rows = [tuple(cell.value for cell in row) for row in ws.iter_rows()]
        wb.close()
        return rows
    except Exception as e:
        log.warning(f'afiliacion: error leyendo hoja {sheet_name} — {e}')
        return None


def _extract_df_from_zip(data: bytes) -> pl.DataFrame | None:
    """Extract CoberturaMunicipio sheet from the MinSalud ZIP into a Polars DataFrame."""
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as z:
            xlsx_names = [n for n in z.namelist() if n.lower().endswith(('.xlsx', '.xls'))]
            if not xlsx_names:
                log.warning('afiliacion: no hay Excel en el ZIP.')
                return None
            with z.open(xlsx_names[0]) as f:
                xlsx_bytes = f.read()
    except Exception as e:
        log.warning(f'afiliacion: error abriendo ZIP — {e}')
        return None

    preferred = ['CoberturaMunicipio', 'EPSMunicipio', 'EPS']
    for sheet in preferred:
        rows = _read_sheet_openpyxl(xlsx_bytes, sheet)
        if not rows:
            continue

        # Find the real header row: first row with ≥3 non-null string values
        header_idx = 0
        for i, row in enumerate(rows[:10]):
            non_null = [v for v in row if v is not None and str(v).strip()]
            if len(non_null) >= 3:
                header_idx = i
                break

        col_names = [str(v).strip() if v is not None else f'col_{j}'
                     for j, v in enumerate(rows[header_idx])]

        # Deduplicate column names
        seen: dict = {}
        deduped = []
        for c in col_names:
            if c not in seen:
                seen[c] = 0
                deduped.append(c)
            else:
                seen[c] += 1
                deduped.append(f'{c}.{seen[c]}')

        data_rows = rows[header_idx + 1:]
        n_cols = len(deduped)
        cols_dict = {
            deduped[j]: [str(row[j]).strip() if row[j] is not None else None
                         for row in data_rows if len(row) > j]
            for j in range(n_cols)
        }
        # Pad shorter columns
        max_len = max((len(v) for v in cols_dict.values()), default=0)
        for k in cols_dict:
            cols_dict[k] += [None] * (max_len - len(cols_dict[k]))

        df = pl.DataFrame(cols_dict)
        if df.is_empty():
            continue

        log.info(f'afiliacion: hoja [{sheet}] — {len(df):,} filas')
        return df

    log.warning('afiliacion: no se encontró ninguna hoja válida.')
    return None


def _normalize_df(df: pl.DataFrame, year: int, month: int) -> pl.DataFrame:
    df = df.rename({c: c.strip() for c in df.columns})
    df = df.rename({k: v for k, v in RENAME_MAP.items() if k in df.columns})

    df = df.with_columns([
        pl.lit(year).cast(pl.Int32).alias('anio'),
        pl.lit(MESES_ES[month].upper()).alias('mes'),
    ])

    for col in ('afiliados_total', 'afiliados_contributivo', 'afiliados_subsidiado', 'afiliados'):
        if col in df.columns:
            df = df.with_columns(
                pl.col(col).str.replace_all(r'[,.]', '').cast(pl.Int64, strict=False)
            )

    df = df.filter(~pl.all_horizontal(pl.all().is_null()))
    return df


def _register_duckdb(parquet_dir: Path, db_path: Path) -> None:
    glob = f'{parquet_dir.as_posix()}/afiliacion_*.parquet'
    con = duckdb.connect(str(db_path))
    con.execute('DROP TABLE IF EXISTS afiliacion')
    con.execute(f"CREATE TABLE afiliacion AS SELECT * FROM read_parquet('{glob}')")
    n = con.execute('SELECT COUNT(*) FROM afiliacion').fetchone()[0]
    con.close()
    print(f'  DuckDB [afiliacion]: {n:,} filas')


def run(base_dir: str = None, year: int = None, month: int = None) -> None:
    base_dir = Path(base_dir) if base_dir else Path(__file__).resolve().parent.parent
    parquet_dir = base_dir / 'data' / 'parquet'
    db_path = base_dir / 'data' / 'informe_regional.duckdb'
    parquet_dir.mkdir(parents=True, exist_ok=True)

    now = datetime.now()
    if year is None or month is None:
        if now.month == 1:
            year, month = now.year - 1, 12
        else:
            year, month = now.year, now.month - 1

    stamp = f'{year}_{month:02d}'
    parquet_path = parquet_dir / f'afiliacion_{stamp}.parquet'

    if parquet_path.exists():
        print(f'afiliacion_scraper: parquet existente {parquet_path}')
        _register_duckdb(parquet_dir, db_path)
        return

    # --- Migrate existing CSV if present ---
    raw_dir = base_dir / 'data' / 'raw'
    csv_path = raw_dir / f'afiliacion_{stamp}.csv'
    if csv_path.exists():
        print(f'afiliacion_scraper: convirtiendo CSV existente a Parquet')
        df = pl.read_csv(csv_path, separator='|', infer_schema_length=0, ignore_errors=True)
        df = df.rename({c: c.strip() for c in df.columns})
        df.write_parquet(parquet_path, compression='zstd')
        print(f'  -> {parquet_path} ({len(df):,} filas)')
        _register_duckdb(parquet_dir, db_path)
        print('afiliacion_scraper OK (desde CSV migrado)')
        return

    # --- Check for ZIP in project root (manually downloaded) ---
    # Only match ZIPs that contain both the month name and year — avoids picking up
    # a different month's ZIP sitting in the project root.
    mes = MESES_ES[month]
    root_zips = [p for p in base_dir.glob('*.zip')
                 if mes in p.name.lower() and str(year) in p.name]

    zip_data = None
    if root_zips:
        log.info(f'afiliacion: usando ZIP existente {root_zips[0]}')
        zip_data = root_zips[0].read_bytes()
        m = re.search(r'(\w+)-(\d{4})\.zip', root_zips[0].name)
        if m:
            mes_str, yr_str = m.group(1).lower(), m.group(2)
            for k, v in MESES_ES.items():
                if v == mes_str:
                    month, year = k, int(yr_str)
                    break
            stamp = f'{year}_{month:02d}'
            parquet_path = parquet_dir / f'afiliacion_{stamp}.parquet'
            if parquet_path.exists():
                print(f'afiliacion_scraper: parquet existente {parquet_path}')
                _register_duckdb(parquet_dir, db_path)
                return
    else:
        url = _build_url(year, month)
        print(f'afiliacion_scraper: descargando {url}')
        zip_data = _download_zip(url)

        if zip_data is None:
            # Fallback: try adjacent months but keep original stamp so data is stored
            # under the requested period, not the fallback period.
            for delta in (-1, 1, -2):
                alt_month = month + delta
                alt_year = year
                if alt_month < 1:
                    alt_month += 12; alt_year -= 1
                elif alt_month > 12:
                    alt_month -= 12; alt_year += 1
                zip_data = _download_zip(_build_url(alt_year, alt_month))
                if zip_data:
                    log.warning(
                        f'afiliacion: usando datos de {alt_year}-{alt_month:02d} '
                        f'como aproximación de {year}-{month:02d}'
                    )
                    break

    if zip_data is None:
        print('afiliacion_scraper: ADVERTENCIA — no se pudo obtener datos.')
        return

    df = _extract_df_from_zip(zip_data)
    if df is None:
        print('afiliacion_scraper: ADVERTENCIA — no se pudo leer Excel del ZIP.')
        return

    df = _normalize_df(df, year, month)
    df.write_parquet(parquet_path, compression='zstd')
    print(f'  -> {parquet_path} ({len(df):,} filas)')

    _register_duckdb(parquet_dir, db_path)
    print('afiliacion_scraper OK')


if __name__ == '__main__':
    logging.basicConfig(level=logging.INFO)
    run()
